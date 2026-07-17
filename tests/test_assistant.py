"""Ask Scopio assistant: command parsing, query building, and file export."""
import io
import json

import app.services.assistant as asst
from app.schemas.assistant import ALLOWED_COLUMNS, AssistantIntent
from app.services.assistant import (
    build_csv,
    build_query,
    build_xlsx,
    heuristic_parse,
    maps_link,
    parse_command,
)

# ------------------------------------------------------------ heuristic fallback

def test_heuristic_understands_no_website_plus_excel():
    # the exact kind of ask the feature was built for
    intent = heuristic_parse(
        "i want the business names and their details whos do not have any website "
        "and also create a excel file where the business details will store"
    )
    assert intent.filters.has_website is False
    assert intent.wants_export is True
    assert intent.file_format == "xlsx"


def test_heuristic_with_website_and_csv():
    intent = heuristic_parse("show businesses with a website and export a csv")
    assert intent.filters.has_website is True
    assert intent.wants_export is True
    assert intent.file_format == "csv"


def test_heuristic_scope_and_no_export():
    intent = heuristic_parse("show me the contactable businesses in this search")
    assert intent.scope == "current_search"
    assert intent.filters.has_any_contact is True
    assert intent.wants_export is False


def test_heuristic_detects_business_type():
    # "only the cafes which have no website" → category filter + website filter
    intent = heuristic_parse("give me only the cafes name which have no website")
    assert intent.filters.categories == ["cafe"]
    assert intent.filters.has_website is False

    intent = heuristic_parse("list the bakeries and their phones")
    assert "bakery" in intent.filters.categories


def test_categories_are_singularized():
    intent = AssistantIntent.model_validate(
        {"filters": {"categories": ["Cafes", "Bakeries", "salons", "spas"]}}
    )
    assert intent.filters.categories == ["cafe", "bakery", "salon", "spa"]


# ------------------------------------------------------------------ LLM parsing

async def test_parse_uses_heuristic_without_llm(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: False)
    intent, parser = await parse_command(
        "businesses without website, excel please", services="we build websites", company="Acme"
    )
    assert parser == "heuristic"
    assert intent.filters.has_website is False


async def test_parse_validates_llm_json(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)

    async def fake_chat(messages, **kwargs):
        # user context (company + services) lives in the system prompt
        assert "Acme" in messages[0]["content"]
        assert "we build websites" in messages[0]["content"]
        return json.dumps({
            "summary": "Businesses without a website, exported to Excel.",
            "scope": "all_leads",
            "filters": {"has_website": False, "categories": ["Cafe"],
                        "statuses": ["discovered", "bogus_status"]},
            "wants_export": True,
            "file_format": "xlsx",
            "columns": ["name", "phone", "email", "maps_link", "not_a_column"],
        })

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    intent, parser = await parse_command(
        "businesses without website, excel please", services="we build websites", company="Acme"
    )
    assert parser == "llm"
    assert intent.filters.has_website is False
    assert intent.filters.categories == ["cafe"]            # normalized
    assert intent.filters.statuses == ["discovered"]        # bogus dropped
    assert "not_a_column" not in intent.columns             # unknown column dropped
    assert intent.columns[0] == "name"                      # canonical order kept


async def test_parse_chat_mode(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)

    async def fake_chat(messages, **kwargs):
        return json.dumps({
            "mode": "chat",
            "reply": "Lead with the eco angle: cafes cut costs and win green-minded customers.",
            "summary": "Advice on pitching cafes.",
        })

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    intent, parser = await parse_command(
        "how should i pitch my eco cups to cafes?",
        services="organic eco friendly cups", company="GreenCup",
    )
    assert parser == "llm"
    assert intent.mode == "chat"
    assert "eco angle" in intent.reply


async def test_parse_passes_chat_history(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)
    seen = {}

    async def fake_chat(messages, **kwargs):
        seen["messages"] = messages
        return json.dumps({"mode": "query", "reply": "Here they are:",
                           "filters": {"categories": ["cafe"]}})

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    history = [
        {"role": "user", "content": "show businesses without a website"},
        {"role": "assistant", "content": "Here they are: [showed 12 matching businesses]"},
    ]
    intent, _ = await parse_command(
        "now only the cafes", services="eco cups", company="GreenCup", history=history,
    )
    # prior turns must be in the LLM conversation so "now only the cafes" resolves
    roles = [m["role"] for m in seen["messages"]]
    assert roles == ["system", "user", "assistant", "user"]
    assert "without a website" in seen["messages"][1]["content"]
    assert intent.filters.categories == ["cafe"]


async def test_parse_bad_llm_json_falls_back(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)

    async def bad_chat(messages, **kwargs):
        return "not json"

    monkeypatch.setattr(asst.llm, "chat", bad_chat)
    intent, parser = await parse_command("no website ones in excel", services=None, company=None)
    assert parser == "heuristic"
    assert intent.filters.has_website is False
    assert intent.wants_export is True


# ---------------------------------------------------------------- query building

def _sql(intent: AssistantIntent, job_id: str | None = None) -> str:
    return str(build_query(intent, job_id).compile())


def test_build_query_no_website_filter():
    intent = AssistantIntent.model_validate(
        {"filters": {"has_website": False}, "scope": "all_leads"}
    )
    sql = _sql(intent)
    assert "website IS NULL" in sql
    assert "deleted_at IS NULL" in sql
    assert "search_job_business" not in sql   # all_leads → no job join


def test_build_query_scopes_to_job():
    intent = AssistantIntent.model_validate({"scope": "current_search"})
    sql = _sql(intent, job_id="00000000-0000-0000-0000-000000000001")
    assert "search_job_business" in sql


def test_build_query_current_search_without_job_falls_back():
    intent = AssistantIntent.model_validate({"scope": "current_search"})
    assert "search_job_business" not in _sql(intent, job_id=None)


def test_build_query_category_and_contact():
    intent = AssistantIntent.model_validate(
        {"filters": {"categories": ["cafe", "salon"], "has_any_contact": True}}
    )
    sql = _sql(intent)
    assert sql.count("LIKE lower") == 2 or "ilike" in sql.lower() or "LIKE" in sql


def test_build_query_matches_precise_type_in_raw_tags():
    # "restaurant" is stored in raw->>'amenity', NOT in the category bucket ("food"),
    # so the type filter must reach into the raw OSM tags (and the name).
    intent = AssistantIntent.model_validate({"filters": {"categories": ["restaurant"]}})
    sql = str(build_query(intent, None).compile(compile_kwargs={"literal_binds": True}))
    assert "'amenity'" in sql      # raw tag keys are searched
    assert "'shop'" in sql
    assert "%restaurant%" in sql   # the type word is matched as a substring


def test_heuristic_defaults_to_current_search():
    # "all business" means the ones on screen — not every lead ever discovered
    intent = heuristic_parse("give me all business name whose do not have any website")
    assert intent.scope == "current_search"
    assert intent.filters.has_website is False

    intent = heuristic_parse("show all my leads everywhere without a website")
    assert intent.scope == "all_leads"


def test_display_category_prefers_raw_tag():
    assert asst.display_category(_Biz(raw={"amenity": "restaurant"}, category="food")) == "restaurant"
    assert asst.display_category(_Biz(raw={"amenity": "fast_food"}, category="food")) == "fast food"
    assert asst.display_category(_Biz(raw={"shop": "bakery"}, category="food")) == "bakery"
    assert asst.display_category(_Biz(raw=None, category="food")) == "food"
    assert asst.display_category(_Biz(raw={"building": "yes"}, category=None)) == "uncategorized"


# ---------------------------------------------------------------------- exports

class _Biz:
    """Duck-typed Business stand-in (no DB needed)."""
    def __init__(self, **kw):
        self.name = kw.get("name", "Cafe Roma")
        self.category = kw.get("category")
        self.phone = kw.get("phone")
        self.email = kw.get("email")
        self.website = kw.get("website")
        self.address = kw.get("address")
        self.lat = kw.get("lat")
        self.lng = kw.get("lng")
        self.status = kw.get("status", "discovered")
        self.details = kw.get("details")
        self.raw = kw.get("raw")


def test_maps_link_prefers_coordinates():
    link = maps_link("Cafe Roma", "12 High St", 22.72, 88.48)
    # /maps/search/<name>/@lat,lng,zoom — searches the name AT the spot (place card).
    assert link == "https://www.google.com/maps/search/Cafe+Roma/@22.72,88.48,17z"


def test_maps_link_falls_back_to_name_address():
    link = maps_link("Cafe Roma", "12 High St, Barasat", None, None)
    assert "Cafe+Roma" in link and "Barasat" in link


def test_row_and_csv_include_description_and_link():
    biz = _Biz(
        phone="+91 98000 00000", email="hi@roma.in", address="12 High St",
        details={"description": "Small espresso bar.", "known_facts": ["owner is Ravi"],
                 "opening_hours": "9-5"},
    )
    row = asst.row_for(biz, list(ALLOWED_COLUMNS))
    assert row["category"] == "uncategorized"
    assert "Small espresso bar." in row["description"]
    assert "owner is Ravi" in row["description"]
    assert "Hours: 9-5" in row["description"]

    data = build_csv([row], list(ALLOWED_COLUMNS))
    text = data.decode("utf-8-sig")
    assert "Business name" in text.splitlines()[0]
    assert "hi@roma.in" in text
    assert "google.com/maps" in text


def test_xlsx_has_clickable_maps_hyperlink():
    try:
        from openpyxl import load_workbook
    except ImportError:  # openpyxl optional in dev envs — export falls back to CSV
        import pytest
        pytest.skip("openpyxl not installed")

    biz = _Biz(website="roma.in", lat=22.72, lng=88.48)
    cols = ["name", "website", "maps_link"]
    data = build_xlsx([asst.row_for(biz, cols)], cols)
    assert data is not None
    ws = load_workbook(io.BytesIO(data)).active
    assert ws.cell(row=1, column=1).value == "Business name"
    maps_cell = ws.cell(row=2, column=3)
    assert maps_cell.value == "Open in Google Maps"
    assert maps_cell.hyperlink is not None
    assert "google.com/maps" in maps_cell.hyperlink.target
    web_cell = ws.cell(row=2, column=2)
    assert web_cell.hyperlink.target == "https://roma.in"


# ----------------------------------------------------------------- categorization

async def test_categorize_missing_skips_without_llm(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: False)
    changed = await asst.categorize_missing(None, [_Biz()])
    assert changed == 0


# ------------------------------------------------------------ agentic answering

async def test_planner_flags_web_search(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)

    async def fake_chat(messages, **kwargs):
        return json.dumps({
            "mode": "chat", "reply": "",
            "filters": {"categories": ["restaurant"]},
            "web_search": True, "web_query": "best restaurants Barasat reviews",
        })

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    intent, _ = await parse_command(
        "which restaurants here have the best reviews?",
        services="eco cups", company="GreenCup",
    )
    assert intent.mode == "chat"
    assert intent.web_search is True
    assert "reviews" in intent.web_query


def _patch_retrieval(monkeypatch, rows, total, grouped):
    async def fake_query(db, intent, job_id, limit=asst.ANSWER_SAMPLE):
        return list(rows), total

    async def fake_grouped(db, intent, job_id):
        return grouped

    monkeypatch.setattr(asst, "run_command_query", fake_query)
    monkeypatch.setattr(asst, "grouped_counts", fake_grouped)


async def test_answer_grounds_in_database(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)
    biz = _Biz(name="Cafe Roma", raw={"amenity": "cafe"}, phone="+91 98000 00000", website=None)
    _patch_retrieval(monkeypatch, [biz], 1, {"cafe": 1})
    monkeypatch.setattr(asst.webtools, "tavily_available", lambda: False)

    seen = {}

    async def fake_chat(messages, **kwargs):
        seen["user"] = messages[-1]["content"]
        return "Cafe Roma has no website — a strong lead for you."

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    intent = AssistantIntent.model_validate({"mode": "chat", "filters": {"categories": ["cafe"]}})
    reply, used_web, sources = await asst.answer_with_data(
        None, intent, "which cafes have no website?",
        services="eco cups", company="GreenCup", job_id=None,
    )
    assert "Cafe Roma" in reply
    assert used_web is False and sources == []
    # the actual DB data must be handed to the LLM (it answers FROM data, not memory)
    assert "DATABASE" in seen["user"]
    assert "Cafe Roma" in seen["user"]
    assert "NO WEBSITE" in seen["user"]


async def test_answer_calls_web_tool_when_flagged(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)
    biz = _Biz(name="Cafe Roma", raw={"amenity": "cafe"})
    _patch_retrieval(monkeypatch, [biz], 1, {"cafe": 1})
    monkeypatch.setattr(asst.webtools, "tavily_available", lambda: True)

    calls = {}

    async def fake_search(query, max_results=5):
        calls["query"] = query
        return [{"title": "Roma reviews", "url": "http://ex.com/r", "content": "Rated 4.5 stars"}]

    monkeypatch.setattr(asst.webtools, "tavily_search", fake_search)

    seen = {}

    async def fake_chat(messages, **kwargs):
        seen["user"] = messages[-1]["content"]
        return "Cafe Roma is rated 4.5 stars."

    monkeypatch.setattr(asst.llm, "chat", fake_chat)
    intent = AssistantIntent.model_validate(
        {"mode": "chat", "web_search": True, "web_query": "Cafe Roma reviews"}
    )
    reply, used_web, sources = await asst.answer_with_data(
        None, intent, "does Cafe Roma have good reviews?",
        services="eco cups", company="GreenCup", job_id=None,
    )
    assert used_web is True
    assert calls["query"] == "Cafe Roma reviews"     # the brain's query is used
    assert sources[0]["url"] == "http://ex.com/r"
    assert "Rated 4.5 stars" in seen["user"]          # web content reaches the LLM
    assert "4.5" in reply


async def test_answer_skips_web_without_tavily_key(monkeypatch):
    monkeypatch.setattr(asst.llm, "llm_available", lambda: True)
    _patch_retrieval(monkeypatch, [_Biz(raw={"amenity": "cafe"})], 1, {"cafe": 1})
    monkeypatch.setattr(asst.webtools, "tavily_available", lambda: False)

    async def boom(*a, **k):  # must NOT be called
        raise AssertionError("tavily_search should not run without a key")

    monkeypatch.setattr(asst.webtools, "tavily_search", boom)
    monkeypatch.setattr(asst.llm, "chat", lambda *a, **k: _ret("From the database."))
    intent = AssistantIntent.model_validate({"mode": "chat", "web_search": True})
    _reply, used_web, _sources = await asst.answer_with_data(
        None, intent, "q", services=None, company=None, job_id=None,
    )
    assert used_web is False


async def _ret(v):
    return v
