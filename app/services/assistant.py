"""The "Ask Scopio" assistant: an agentic mini-chatbot over the user's leads.

The LLM is the BRAIN, not the answer source. For each message it PLANS (parse_command):
either a leads query (filter/list/export) or a question to answer. Answers are grounded
in the user's DATABASE of discovered businesses — the ones stored when they searched —
and, when the database can't satisfy the question, the LLM calls the WEB tool (Tavily)
and answers from what comes back (answer_with_data). Plain SQLAlchemy does the retrieval;
the LLM never touches the DB directly.

Degrades gracefully: with no LLM key a keyword heuristic still handles list/filter/export;
without a Tavily key the web step is simply skipped and answers use the DB alone.
"""
import csv
import io
import json
import logging
import re
from urllib.parse import quote_plus

from sqlalchemy import Select, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import llm
from app.models.business import Business
from app.models.search_job_business import SearchJobBusiness
from app.schemas.assistant import AssistantFilters, AssistantIntent
from app.services.deepagent import tools as webtools

logger = logging.getLogger(__name__)

MAX_RESULTS = 500          # rows returned to the UI table per command (drill-in shows the rest)
EXPORT_MAX = 50000         # rows written to an exported file (effectively "all")
ANSWER_SAMPLE = 80         # DB rows fed to the answer-synthesis LLM call
CATEGORIZE_CAP = 40        # max businesses AI-categorized per command (token budget)


# --------------------------------------------------------------------------- parse

_SYSTEM = (
    "You are Scopio's assistant — the BRAIN of a B2B lead tool. You do NOT answer "
    "from your own knowledge; you answer from the user's DATABASE of discovered "
    "businesses (stored when they searched) and, when that's not enough, from the "
    "WEB tool. You know the user's own business (company + what they sell), so every "
    "answer is tailored to THEIR business.\n"
    "You are a PLANNER. For each message, decide what to do:\n"
    '(a) mode "query" — the user wants to SEE/FILTER/EXPORT their leads as a list '
    '("show cafes", "all restaurants without a website", "export an excel"). Fill '
    "in the filters; the tool lists the matching businesses.\n"
    '(b) mode "chat" — the user asks a QUESTION or wants ADVICE ("which area has the '
    'most restaurants?", "who should I contact first?", "does X have good reviews?", '
    '"how do I pitch a bakery?"). Set filters to narrow which leads are relevant to '
    "the question (e.g. category=restaurant), and decide whether the DATABASE alone "
    "can answer it. If the answer needs facts NOT in the database (reviews, ratings, "
    "recent news, who owns it, general market info), set web_search=true and give a "
    "web_query. The tool will fetch the data and you'll answer from it next.\n"
    "Never fabricate business names, phones or facts — those come from the data."
)

_INSTRUCTIONS = (
    "Return ONLY a JSON object with these keys:\n"
    '  "mode": "query" (list/filter/export leads) or "chat" (answer a question / '
    "give advice).\n"
    '  "reply": for query mode, one short lead-in ("Here are the restaurants:"). '
    "For chat mode, leave it SHORT or empty — the real answer is written afterwards "
    "from the fetched data.\n"
    '  "summary": one short sentence restating what they asked for.\n'
    '  "scope": "current_search" (businesses of the search on screen — DEFAULT when '
    'a search is open) or "all_leads" (every lead ever discovered — only when they '
    "clearly ask for all/everywhere/entire).\n"
    '  "filters": object — has_website, has_phone, has_email, has_any_contact (each '
    "true/false/null; null = no filter; e.g. WITHOUT a website → has_website=false), "
    '"categories" (array of SINGULAR lowercase business-type words, e.g. "only the '
    'cafes" → ["cafe"]; [] = all types), "statuses" (from: discovered, contacted, '
    "interested, callback_scheduled, not_interested, do_not_contact; [] = all), "
    '"name_contains" (string or null).\n'
    '  "wants_export": true if they want a file (Excel/CSV/spreadsheet/download).\n'
    '  "file_format": "xlsx" or "csv" (default "xlsx").\n'
    '  "columns": array from [name, category, phone, email, website, address, '
    "description, status, maps_link]; if they ask for ONLY names use [\"name\"]; "
    "default = all.\n"
    '  "web_search": true ONLY in chat mode when the database cannot answer and the '
    "web is needed. Otherwise false.\n"
    '  "web_query": the search query to run when web_search is true (include the '
    "business name + location for a specific business), else null.\n"
    "No prose, no markdown — just the JSON object."
)


# The assistant's second pass: write the final answer grounded in fetched data.
_ANSWER_SYSTEM = (
    "You are Scopio's assistant, answering the user's question about THEIR discovered "
    "business leads. Answer ONLY from the DATA provided below (database rows + any web "
    "results) plus the user's own business context — never invent businesses, phone "
    "numbers, or facts. Be specific: name real businesses from the data, give counts, "
    "and make it useful for their sales work. If the data doesn't contain the answer, "
    "say so plainly and suggest what to search or enrich next. Plain text, 2-8 "
    "sentences, no markdown headers. Answer in the user's language."
)


# Business-type words the no-LLM fallback can recognize in a command.
_KNOWN_TYPES = (
    "cafe", "restaurant", "bakery", "bar", "hotel", "spa", "salon", "office",
    "gym", "clinic", "pharmacy", "school", "grocery", "store", "shop", "bank",
)


def heuristic_parse(command: str) -> AssistantIntent:
    """Keyword fallback when no LLM is configured (or it fails). Covers the basics."""
    text = command.lower()
    filters = AssistantFilters()

    # "only the cafes…" → categories=["cafe"] (plurals normalized by the schema)
    def _pat(t: str) -> str:
        return rf"\b{t[:-1]}(y|ies)\b" if t.endswith("y") else rf"\b{t}(s|es)?\b"
    filters.categories = [t for t in _KNOWN_TYPES if re.search(_pat(t), text)]

    if re.search(r"(no|without|don'?t\s+have|do(es)?\s+not\s+have|missing|lack(ing)?)"
                 r"[\w\s]{0,20}\bweb\s?site", text):
        filters.has_website = False
    elif re.search(r"(with|having|has|have)\s+(a\s+|an?y\s+)?web\s?site", text):
        filters.has_website = True

    if re.search(r"contactable|can\s+contact|with\s+contact", text):
        filters.has_any_contact = True
    if re.search(r"(no|without|missing)\s+(a\s+)?phone", text):
        filters.has_phone = False
    if re.search(r"(no|without|missing)\s+(an?\s+)?e-?mail", text):
        filters.has_email = False

    wants_export = bool(re.search(r"excel|xlsx|spread\s?sheet|csv|download|export"
                                  r"|make\s+(me\s+)?a\s+file|save\s+(it|them)", text))
    file_format = "csv" if ("csv" in text and "excel" not in text and "xlsx" not in text) else "xlsx"
    # Default to what's on screen; only go tenant-wide when they clearly ask for it.
    scope = ("all_leads" if re.search(
        r"\ball\s+(my\s+)?(leads|searches|areas|places)\b|\beverywhere\b|\bentire\b|\bever\b",
        text) else "current_search")

    summary = "Show matching businesses" + (" and export a file" if wants_export else "")
    return AssistantIntent(
        mode="query", reply="Here's what matches your request:",
        summary=summary, scope=scope, filters=filters,
        wants_export=wants_export, file_format=file_format,
    )


async def parse_command(
    command: str,
    *,
    services: str | None,
    company: str | None,
    history: list[dict] | None = None,
    data_context: str | None = None,
) -> tuple[AssistantIntent, str]:
    """Free text (+ chat history + a snapshot of the visible leads) → intent."""
    if not llm.llm_available():
        return heuristic_parse(command), "heuristic"
    # The user's business context makes "my kind of customers" style phrasing resolve.
    context = (services or "").strip()
    if len(context) > 1500:                       # services can be ~20k words now
        context = context[:1500] + "…"
    system = (
        f"{_SYSTEM}\n\nUser's company: {company or 'unknown'}\n"
        f"What the user's business offers:\n{context or '(not provided)'}"
    )
    if data_context:
        # A live snapshot of the leads on screen: lets chat mode answer questions
        # about specific businesses ("what do you think of X?", "who should I call
        # first?") with real data instead of guessing.
        system += f"\n\n{data_context}"
    messages: list[dict] = [{"role": "system", "content": system}]
    # Chat memory: prior turns so follow-ups ("only the cafes", "now as excel") resolve.
    for turn in (history or [])[-10:]:
        role = turn.get("role")
        content = str(turn.get("content") or "")[:2000]
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": f"{command}\n\n{_INSTRUCTIONS}"})
    try:
        content = await llm.chat(messages, json_mode=True, max_tokens=900)
        intent = AssistantIntent.model_validate(json.loads(content))
        return intent, "llm"
    except Exception as exc:  # noqa: BLE001 — parser must never take the feature down
        logger.warning("assistant parse failed, using heuristic: %s", exc)
        return heuristic_parse(command), "heuristic"


# --------------------------------------------------------------------------- query

# Raw OSM tag keys that carry the PRECISE business type ("restaurant", "bakery"…).
# `business.category` only stores a broad bucket (food/retail/finance/…), so type
# words from the user must be matched against these tags too.
_TYPE_TAG_KEYS = ("amenity", "shop", "office", "craft", "tourism", "leisure",
                  "healthcare", "cuisine")


def _empty(col) -> object:
    """NULL or '' — discovery sometimes stores empty strings."""
    return or_(col.is_(None), col == "")


def _present(col) -> object:
    return func.coalesce(col, "") != ""


def _category_clause(term: str) -> object:
    """Match a business-type word against the bucket, the raw OSM tags and the name.

    "restaurant" lives in raw->>'amenity', "bakery" in raw->>'shop'; the `category`
    column only says "food" — so all three must be checked.
    """
    like = f"%{term}%"
    clauses = [Business.category.ilike(like), Business.name.ilike(like)]
    clauses += [Business.raw[k].astext.ilike(like) for k in _TYPE_TAG_KEYS]
    return or_(*clauses)


def _conditions(intent: AssistantIntent) -> list:
    f = intent.filters
    conds = [Business.deleted_at.is_(None)]

    if f.has_website is True:
        conds.append(_present(Business.website))
    elif f.has_website is False:
        conds.append(_empty(Business.website))
    if f.has_phone is True:
        conds.append(_present(Business.phone))
    elif f.has_phone is False:
        conds.append(_empty(Business.phone))
    if f.has_email is True:
        conds.append(_present(Business.email))
    elif f.has_email is False:
        conds.append(_empty(Business.email))
    if f.has_any_contact is True:
        conds.append(or_(_present(Business.phone), _present(Business.email)))
    elif f.has_any_contact is False:
        conds.append(_empty(Business.phone))
        conds.append(_empty(Business.email))
    if f.categories:
        conds.append(or_(*[_category_clause(c) for c in f.categories]))
    if f.statuses:
        conds.append(Business.status.in_(f.statuses))
    if f.name_contains:
        conds.append(Business.name.ilike(f"%{f.name_contains}%"))
    return conds


def _scoped(q, intent: AssistantIntent, job_id: str | None):
    if intent.scope == "current_search" and job_id:
        q = q.join(SearchJobBusiness, SearchJobBusiness.business_id == Business.id).where(
            SearchJobBusiness.search_job_id == job_id
        )
    return q


# The precise on-display type: raw OSM tag value when present, else the bucket.
_display_type_sql = func.coalesce(
    Business.raw["amenity"].astext, Business.raw["shop"].astext,
    Business.raw["craft"].astext, Business.raw["tourism"].astext,
    Business.category, "uncategorized",
)


def build_query(intent: AssistantIntent, job_id: str | None) -> Select:
    """Intent → SELECT over the tenant's businesses (RLS scopes the tenant)."""
    q = _scoped(select(Business), intent, job_id).where(*_conditions(intent))
    return q.order_by(_display_type_sql, Business.name)


async def run_command_query(
    db: AsyncSession, intent: AssistantIntent, job_id: str | None, limit: int = MAX_RESULTS
) -> tuple[list[Business], int]:
    base = build_query(intent, job_id)
    total = (await db.execute(
        select(func.count()).select_from(base.order_by(None).subquery())
    )).scalar_one()
    rows = (await db.execute(base.limit(limit))).scalars().all()
    return list(rows), int(total)


# Compare a chip label ("fast food") back to the stored type value ("fast_food").
_display_label_sql = func.replace(_display_type_sql, "_", " ")


async def list_by_category(
    db: AsyncSession, intent: AssistantIntent, job_id: str | None,
    category: str, limit: int, offset: int,
) -> tuple[list[Business], int]:
    """Businesses of one display-type within the current result set (for drill-in)."""
    conds = _conditions(intent)
    label = category.strip().replace("_", " ").lower()
    match = func.lower(_display_label_sql) == label
    q = _scoped(select(Business), intent, job_id).where(*conds, match)
    total = (await db.execute(
        select(func.count()).select_from(q.order_by(None).subquery())
    )).scalar_one()
    rows = (await db.execute(
        q.order_by(Business.name).limit(limit).offset(offset)
    )).scalars().all()
    return list(rows), int(total)


# --------------------------------------------------------------- agentic answering

def _biz_line(biz: Business) -> str:
    """One compact, factual line about a business for the answer-synthesis prompt."""
    bits = [biz.name, display_category(biz)]
    if biz.address:
        bits.append(biz.address)
    bits.append(f"phone {biz.phone}" if biz.phone else "no phone")
    bits.append(f"email {biz.email}" if biz.email else "no email")
    bits.append(f"website {biz.website}" if biz.website else "NO WEBSITE")
    if biz.status != "discovered":
        bits.append(f"status {biz.status}")
    desc = (biz.details or {}).get("description")
    if desc:
        bits.append(str(desc)[:120])
    return " | ".join(bits)


async def answer_with_data(
    db: AsyncSession,
    intent: AssistantIntent,
    question: str,
    *,
    services: str | None,
    company: str | None,
    job_id: str | None,
    history: list[dict] | None = None,
) -> tuple[str, bool, list[dict]]:
    """Answer a question grounded in the DB (and the web tool if the plan asked for it).

    Returns (reply, used_web, sources). The LLM is the brain: it already decided the
    filters + whether to web-search in `intent`; here we RETRIEVE, optionally call the
    web tool, then let it write the final answer from that data.
    """
    # 1. Retrieve the relevant leads from the database (per the planned filters).
    rows, total = await run_command_query(db, intent, job_id, limit=ANSWER_SAMPLE)
    by_type = await grouped_counts(db, intent, job_id)

    data_parts = [
        f"DATABASE — {total} matching business(es)"
        + (f"; by type: {', '.join(f'{t}: {n}' for t, n in list(by_type.items())[:15])}" if by_type else "")
        + (f" (showing first {len(rows)})" if total > len(rows) else "") + ":"
    ]
    data_parts += [f"- {_biz_line(b)}" for b in rows] or ["(no businesses match)"]

    # 2. Web tool — only if the brain flagged it AND a key is configured.
    used_web, sources = False, []
    if intent.web_search and webtools.tavily_available():
        query = (intent.web_query or question)[:300]
        results = await webtools.tavily_search(query, max_results=5)
        if results:
            used_web = True
            sources = [{"title": r["title"], "url": r["url"]} for r in results if r.get("url")]
            data_parts.append("\nWEB RESULTS:")
            data_parts += [
                f"- [{r['title']}] {r['content'][:400]} ({r['url']})"
                for r in results if r.get("content")
            ]

    # 3. Synthesize the final answer from the gathered data.
    context = (services or "").strip()[:1500]
    system = (
        f"{_ANSWER_SYSTEM}\n\nUser's company: {company or 'unknown'}\n"
        f"What the user's business offers:\n{context or '(not provided)'}"
    )
    messages: list[dict] = [{"role": "system", "content": system}]
    for turn in (history or [])[-6:]:
        role, content = turn.get("role"), str(turn.get("content") or "")[:1500]
        if role in ("user", "assistant") and content:
            messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content":
                     f"Question: {question}\n\nDATA:\n" + "\n".join(data_parts)})
    try:
        reply = await llm.chat(messages, max_tokens=700, temperature=0.4)
        return reply.strip(), used_web, sources
    except Exception as exc:  # noqa: BLE001 — never crash the chat
        logger.warning("assistant answer synthesis failed: %s", exc)
        fallback = intent.reply or (
            f"I found {total} matching businesses but couldn't compose an answer just "
            "now — try again in a moment."
        )
        return fallback, used_web, sources


async def grouped_counts(
    db: AsyncSession, intent: AssistantIntent, job_id: str | None
) -> dict[str, int]:
    """Type → count over the WHOLE match set (not just the page shown)."""
    q = _scoped(
        select(_display_type_sql, func.count()), intent, job_id
    ).where(*_conditions(intent)).group_by(_display_type_sql).order_by(func.count().desc())
    rows = (await db.execute(q.limit(30))).all()
    return {str(cat).replace("_", " "): int(n) for cat, n in rows}


def display_category(biz: Business) -> str:
    """The precise business type for display: "restaurant", not the bucket "food"."""
    raw = biz.raw or {}
    for k in ("amenity", "shop", "craft", "office", "tourism", "leisure", "healthcare"):
        v = raw.get(k)
        if v and isinstance(v, str) and v not in ("yes", "true", "no"):
            return v.replace("_", " ")
    return biz.category or "uncategorized"


SNAPSHOT_SAMPLE = 60       # businesses listed in the chat context snapshot


async def data_snapshot(db: AsyncSession, job_id: str | None) -> str:
    """A compact, LLM-readable snapshot of the leads the user is looking at.

    Injected into the assistant's system prompt so chat mode can answer questions
    and give advice about the actual businesses on screen (any place, any search).
    """
    intent = AssistantIntent(scope="current_search" if job_id else "all_leads")
    conds = _conditions(intent)

    total = (await db.execute(
        _scoped(select(func.count(Business.id)), intent, job_id).where(*conds)
    )).scalar_one()
    if not total:
        return ("DATA SNAPSHOT: the user has no discovered businesses on screen yet. "
                "Suggest running a search (Find businesses) first.")

    by_type = await grouped_counts(db, intent, job_id)
    no_web = (await db.execute(
        _scoped(select(func.count(Business.id)), intent, job_id)
        .where(*conds, _empty(Business.website))
    )).scalar_one()
    no_contact = (await db.execute(
        _scoped(select(func.count(Business.id)), intent, job_id)
        .where(*conds, _empty(Business.phone), _empty(Business.email))
    )).scalar_one()

    sample = (await db.execute(
        _scoped(select(Business), intent, job_id).where(*conds)
        .order_by(Business.enriched_at.desc().nulls_last(), Business.name)
        .limit(SNAPSHOT_SAMPLE)
    )).scalars().all()

    lines = []
    for b in sample:
        bits = [b.name, display_category(b)]
        bits.append(f"phone:{b.phone}" if b.phone else "no phone")
        bits.append(f"email:{b.email}" if b.email else "no email")
        bits.append("has website" if b.website else "NO WEBSITE")
        if b.status != "discovered":
            bits.append(f"status:{b.status}")
        desc = (b.details or {}).get("description")
        if desc:
            bits.append(str(desc)[:90])
        lines.append(" | ".join(bits))

    types_str = ", ".join(f"{t}: {n}" for t, n in list(by_type.items())[:12])
    header = (
        f"DATA SNAPSHOT — the businesses the user is looking at right now "
        f"({'current search' if job_id else 'all their leads'}):\n"
        f"Total: {total} · without a website: {no_web} · without any contact: {no_contact}\n"
        f"By type: {types_str}\n"
        f"Sample ({len(sample)} of {total}):\n"
    )
    return header + "\n".join(f"- {ln}" for ln in lines)


# ------------------------------------------------------------------- categorization

_CAT_SYSTEM = (
    "You label local businesses with a short lowercase business-type category "
    "(e.g. cafe, restaurant, salon, clinic, pharmacy, grocery, electronics, gym, "
    "school, hotel). One or two words max, no punctuation."
)


async def categorize_missing(db: AsyncSession, businesses: list[Business]) -> int:
    """AI-categorize businesses that have no category yet; persists to the DB.

    Best-effort: skipped without an LLM key; any failure categorizes nothing.
    """
    todo = [b for b in businesses if not (b.category or "").strip()][:CATEGORIZE_CAP]
    if not todo or not llm.llm_available():
        return 0
    numbered = "\n".join(
        f"{i}. {b.name}" + (f" — {b.address}" if b.address else "")
        for i, b in enumerate(todo, 1)
    )
    user = (
        "Categorize each business. Return ONLY a JSON object mapping the number to "
        'the category, e.g. {"1": "cafe", "2": "salon"}. Use "other" if truly unclear.\n\n'
        + numbered
    )
    try:
        content = await llm.chat(
            [{"role": "system", "content": _CAT_SYSTEM}, {"role": "user", "content": user}],
            json_mode=True, max_tokens=800, temperature=0.1,
        )
        mapping = json.loads(content)
        changed = 0
        for i, biz in enumerate(todo, 1):
            label = str(mapping.get(str(i)) or "").strip().lower()[:40]
            if label:
                biz.category = label
                changed += 1
        if changed:
            await db.commit()
        return changed
    except Exception as exc:  # noqa: BLE001 — categorization is a bonus, never a blocker
        logger.warning("assistant categorization failed: %s", exc)
        return 0


# --------------------------------------------------------------------------- output

def maps_link(name: str, address: str | None, lat: float | None, lng: float | None) -> str:
    """A clickable Google Maps link that lands on the business's place card."""
    if lat is not None and lng is not None:
        # Search the name centered on the exact coordinates — Google resolves this to
        # the actual place card (a plain "name @lat,lng" query is NOT understood).
        return f"https://www.google.com/maps/search/{quote_plus(name)}/@{lat},{lng},17z"
    query = f"{name} {address}" if address else name
    return "https://www.google.com/maps/search/?api=1&query=" + quote_plus(query)


def description_of(biz: Business) -> str | None:
    """Best human description we have: AI-enriched summary + learned facts."""
    details = biz.details or {}
    parts = [details.get("description")]
    facts = details.get("known_facts") or []
    if isinstance(facts, list) and facts:
        parts.append("; ".join(str(f) for f in facts[:3]))
    hours = details.get("opening_hours") or (biz.raw or {}).get("opening_hours")
    if hours:
        parts.append(f"Hours: {hours}")
    text = " · ".join(p for p in parts if p)
    return text or None


_HEADERS = {
    "name": "Business name", "category": "Category", "phone": "Phone",
    "email": "Email", "website": "Website", "address": "Address",
    "description": "Description", "status": "Status", "maps_link": "Google Maps",
}


def row_for(biz: Business, columns: list[str]) -> dict[str, str]:
    values = {
        "name": biz.name,
        "category": display_category(biz),
        "phone": biz.phone or "",
        "email": biz.email or "",
        "website": biz.website or "",
        "address": biz.address or "",
        "description": description_of(biz) or "",
        "status": biz.status,
        "maps_link": maps_link(biz.name, biz.address, biz.lat, biz.lng),
    }
    return {c: values[c] for c in columns}


def build_csv(rows: list[dict[str, str]], columns: list[str]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\r\n")
    writer.writerow([_HEADERS[c] for c in columns])
    for row in rows:
        writer.writerow([row[c] for c in columns])
    return buf.getvalue().encode("utf-8-sig")   # BOM so Excel opens UTF-8 correctly


def build_xlsx(rows: list[dict[str, str]], columns: list[str]) -> bytes | None:
    """Real .xlsx with clickable hyperlinks. None if openpyxl isn't installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed — assistant export falls back to CSV")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    bold = Font(bold=True)
    for col_idx, c in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADERS[c])
        cell.font = bold

    link_font = Font(color="0563C1", underline="single")
    for r, row in enumerate(rows, 2):
        for col_idx, c in enumerate(columns, 1):
            value = row[c]
            cell = ws.cell(row=r, column=col_idx)
            if c == "maps_link" and value:
                cell.value = "Open in Google Maps"     # click → place details
                cell.hyperlink = value
                cell.font = link_font
            elif c == "website" and value:
                cell.value = value
                cell.hyperlink = value if "://" in value else "https://" + value
                cell.font = link_font
            else:
                cell.value = value

    widths = {"name": 32, "category": 16, "phone": 18, "email": 28, "website": 30,
              "address": 40, "description": 50, "status": 14, "maps_link": 22}
    for col_idx, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(c, 20)
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
